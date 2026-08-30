#!/usr/bin/env python
"""Build the labelling workbook -- ONE sheet, all jurisdictions, sorted by region."""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "UCC_labelled_v3.xlsx"
HDR = PatternFill("solid", fgColor="1F3864"); BAND = PatternFill("solid", fgColor="EEF2F8")
TODO = PatternFill("solid", fgColor="FFF3CD"); PREM = PatternFill("solid", fgColor="E8F1E4")
THIN = Side(style="thin", color="C8CDD6")

k = pd.read_csv(ROOT / "workbook_key_labelled.csv").fillna("")
wb = Workbook(); ws = wb.active; ws.title = "LABEL THESE"
cols = ["pair_id", "region", "question",
        "A — name", "A — address", "A — city", "A — zip",
        "B — name", "B — address", "B — city", "B — zip", "label", "note"]
ws.append(cols)
for c in ws[1]:
    c.fill = HDR; c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
src = ["pair_id", "jurisdiction", "question", "an", "aa", "ac", "az", "bn", "ba", "bc", "bz"]
for _, r in k.iterrows():
    ws.append([r[c] for c in src] + [r.get("label",""), r.get("note","")])

n = len(k) + 1
for row in range(2, n + 1):
    is_prem = str(ws.cell(row=row, column=3).value).startswith("PREMISES")
    for col in range(8, 12):
        ws.cell(row=row, column=col).fill = PREM if is_prem else BAND
    if is_prem:
        ws.cell(row=row, column=3).fill = PREM
    for col in range(1, 14):
        ws.cell(row=row, column=col).border = Border(bottom=THIN)

dv = DataValidation(type="list",
                    formula1='"SAME,DIFFERENT,ONE-OP,SEPARATE,UNSURE"',
                    allow_blank=True, showDropDown=False)
dv.error = "ENTITY rows: SAME / DIFFERENT / UNSURE.  PREMISES rows: ONE-OP / SEPARATE / UNSURE."
dv.errorTitle = "Not a valid label"
ws.add_data_validation(dv); dv.add(f"L2:L{n}")
ws.conditional_formatting.add(f"A2:M{n}", FormulaRule(formula=['$L2=""'], fill=TODO, stopIfTrue=False))
for col, w in zip("ABCDEFGHIJKLM", [10, 13, 26, 34, 26, 15, 11, 34, 26, 15, 11, 13, 24]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "D2"; ws.auto_filter.ref = f"A1:M{n}"

rs = wb.create_sheet("THE RULE")
L = [
 ("TWO DIFFERENT QUESTIONS. Check the `question` column on every row.", 14, True), ("", 11, False),
 ("1) ENTITY - 'same firm?'   ->  SAME / DIFFERENT / UNSURE", 13, True),
 ("   Read the two NAMES first. ~90% decide in 3 seconds from names alone.", 11, False),
 ("   R0  identical name AND identical address -> SAME", 11, False),
 ("   R1  plainly different firms, no shared address -> DIFFERENT   <- commonest case", 11, False),
 ("   R2  one name is a typo/truncation/abbreviation of the other -> SAME", 11, False),
 ("        AMERICAN NATIONAL BANJ = AMERICAN NATIONAL BANK.  COLO = COLORADO.  Tested BEFORE R6.", 11, False),
 ("   R3  suffix-only difference (LLC vs INC) -> SAME", 11, False),
 ("   R4  same name, same city, different address -> SAME  (moved, or a yard and an office)", 11, False),
 ("   R5  same name, DIFFERENT cities, nothing shared -> DIFFERENT", 11, False),
 ("   R6  DIFFERENT NAMES AT THE SAME ADDRESS -> DIFFERENT", 11, True),
 ("        Registered agents, franchise HQs and family properties park unrelated filers at one address.", 11, False),
 ("        ONE exception: numbered outlets of a chain (BUFFET 103 vs BUFFET 500) = SAME, name says so.", 11, False),
 ("   R7  distinct legal entities of one family -> DIFFERENT (CAT FINANCIAL vs CATERPILLAR INC)", 11, False),
 ("   R8  two DIFFERENT person-names at one address -> DIFFERENT. Same person twice -> SAME.", 11, False),
 ("   R9  successor / renamed entity -> DIFFERENT (we label records as filed, not corporate history)", 11, False),
 ("", 11, False),
 ("2) PREMISES - 'one operation?'   ->  ONE-OP / SEPARATE / UNSURE      (green rows)", 13, True),
 ("   These are DIFFERENT names at the SAME address. You already know they are different FIRMS.", 11, False),
 ("   The question is whether they are ONE OPERATION - one yard, one farm, one family business", 11, False),
 ("   buying equipment under more than one name.", 11, False),
 ("     ONE-OP    -> WOOD DONNA L / WOOD DONALD J at one PO box. One farm, two named borrowers.", 11, False),
 ("                  Also: a person and their company at one address. Related SPEs at one office.", 11, False),
 ("     SEPARATE  -> a registered agent's address, a franchise HQ, an office building, a mall unit.", 11, False),
 ("                  Unrelated businesses that merely share a mailing address.", 11, False),
 ("   Rule of thumb: would ONE lender assessing exposure at this address want BOTH liens? -> ONE-OP.", 11, False),
 ("", 11, False),
 ("SCOPE: construction and heavy machinery only. 1990 onward. Colorado and Connecticut are", 12, True),
 ("SEPARATE REGISTERS and are never compared to each other - the `region` column tells you which.", 11, False),
 ("", 11, False),
 ("DO NOT look anything up. No Google, no Secretary of State search. Outside research makes the", 12, True),
 ("labels unreproducible and is not available at scale, so a reviewer would discount the number.", 11, False),
 ("Yellow = unlabelled. When nothing is yellow you are done.", 11, False),
]
for t, sz, b in L:
    rs.append([t]); rs.cell(row=rs.max_row, column=1).font = Font(size=sz, bold=b)
rs.column_dimensions["A"].width = 112
wb.move_sheet("THE RULE", offset=-1); wb.active = 1
wb.save(OUT)
print(f"wrote {OUT}  ({len(k)} rows)")
print(k.groupby(["jurisdiction", "question"]).size().to_string())
for c in ("sim", "band", "anc", "bnc", "aac", "bac", "asuf", "bsuf"):
    assert c not in cols
print("BLIND CHECK: no similarity, band or normalised-key column. PASS")
