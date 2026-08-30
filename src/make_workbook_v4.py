#!/usr/bin/env python
"""v4 workbook -- one sheet, blank labels, approved fields, sorted by region."""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "UCC_labelling_v4.xlsx"
HDR = PatternFill("solid", fgColor="1F3864"); BAND = PatternFill("solid", fgColor="EEF2F8")
TODO = PatternFill("solid", fgColor="FFF3CD"); PREM = PatternFill("solid", fgColor="E8F1E4")
THIN = Side(style="thin", color="C8CDD6")

k = pd.read_csv(ROOT / "sheet_v4_key.csv").fillna("")
wb = Workbook(); ws = wb.active; ws.title = "LABEL THESE"
cols = ["pair_id", "region", "question",
        "A — name", "A — address", "A — city", "A — state", "A — zip",
        "A — lender", "A — loans", "A — first", "A — last",
        "B — name", "B — address", "B — city", "B — state", "B — zip",
        "B — lender", "B — loans", "B — first", "B — last", "label", "note"]
src = ["pair_id", "region", "question", "an", "aa", "ac", "ast", "az", "al", "aln", "afl", "all_",
       "bn", "ba", "bc", "bst", "bz", "bl", "bln", "bfl", "bll"]
ws.append(cols)
for c in ws[1]:
    c.fill = HDR; c.font = Font(bold=True, color="FFFFFF", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for _, r in k.iterrows():
    ws.append([r.get(c, "") for c in src] + ["", ""])

n = len(k) + 1
for row in range(2, n + 1):
    prem = str(ws.cell(row=row, column=3).value).startswith("PREMISES")
    for col in range(13, 22):
        ws.cell(row=row, column=col).fill = PREM if prem else BAND
    if prem:
        ws.cell(row=row, column=3).fill = PREM
    for col in range(1, 24):
        ws.cell(row=row, column=col).border = Border(bottom=THIN)

dv = DataValidation(type="list", formula1='"SAME,DIFFERENT,ONE-OP,SEPARATE,UNSURE"',
                    allow_blank=True, showDropDown=False)
dv.error = "ENTITY rows: SAME / DIFFERENT / UNSURE.  PREMISES rows: ONE-OP / SEPARATE / UNSURE."
dv.errorTitle = "Not a valid label"
ws.add_data_validation(dv); dv.add(f"V2:V{n}")
ws.conditional_formatting.add(f"A2:W{n}", FormulaRule(formula=['$V2=""'], fill=TODO, stopIfTrue=False))
widths = [9, 7, 24, 30, 24, 14, 7, 10, 26, 7, 7, 7, 30, 24, 14, 7, 10, 26, 7, 7, 7, 12, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
ws.freeze_panes = "D2"; ws.auto_filter.ref = f"A1:W{n}"

rs = wb.create_sheet("SCOPE & RULES")
L = [("SCOPE — heavy construction equipment finance only", 14, True), ("", 10, False),
 ("Every row qualifies by ONE of two routes:", 12, True),
 ("  ROUTE A  the LENDER is a heavy-construction manufacturer, captive or dealer —", 10, False),
 ("           Caterpillar, John Deere Construction, Komatsu, Kubota, CNH, Volvo, Terex, JLG,", 10, False),
 ("           Vermeer, Bobcat, Takeuchi, Liebherr, Wagner Equipment, Faris, Honnen, 4 Rivers.", 10, False),
 ("  ROUTE B  the BORROWER'S OWN NAME carries an equipment or trade word — excavation,", 10, False),
 ("           crane, dozer, grading, paving, asphalt, aggregate, drilling, demolition,", 10, False),
 ("           earthwork, pipeline, heavy haul, and the CONCRETE family (concrete, cement,", 10, False),
 ("           mixer, pumping, shotcrete, precast, rebar) — a concrete outfit runs mixers", 10, False),
 ("           and boom trucks, so it qualifies on its name alone.", 10, False), ("", 10, False),
 ("Every row also has: a named party, a real address (no placeholders), and a loan year.", 11, True),
 ("Lender may be blank — banks will not disclose and you cannot cold-call them anyway.", 10, False),
 ("Regions CO and CT are SEPARATE registers. No pair ever crosses between them.", 11, True), ("", 10, False),
 ("QUESTION 1 — ENTITY 'same firm?'  ->  SAME / DIFFERENT / UNSURE", 13, True),
 ("  Names first. ~90% decide on names alone.", 10, False),
 ("  Identity: name is one string split on spaces, MIDDLE INITIALS IGNORED, and a match", 10, False),
 ("  forwards OR backwards is the same party — HOWARD JOHN F = JOHN HOWARD.", 10, False),
 ("  Company not person: a dash between surnames (Stutsman-Gerbaz); any trade word", 10, False),
 ("  (Hernandez Excavating); possessive-plural (Cohen's, Spencers). O'Brian is a person.", 10, False),
 ("  Entity type is identity: LLC is NOT INC.", 10, False),
 ("  Address formatting is not an address difference: RD=ROAD, ST=STREET, HWY=HIGHWAY,", 10, False),
 ("  S=SOUTH — same address if the street number matches and the abbreviation expands.", 10, False),
 ("  A blank field is not evidence. A ZIP that disagrees while name and street agree is", 10, False),
 ("  a data error — ignore it.", 10, False), ("", 10, False),
 ("QUESTION 2 — PREMISES 'one operation?'  ->  ONE-OP / SEPARATE / UNSURE  (green)", 13, True),
 ("  Different names at the SAME address. They are different FIRMS — the question is", 10, False),
 ("  whether they are one OPERATION.", 10, False),
 ("    ONE-OP    a married couple who co-signed. A person and their company. Related SPEs.", 10, False),
 ("    SEPARATE  registered agent, franchise HQ, office building, mall unit, virtual office.", 10, False),
 ("  TEST: would ONE lender assessing exposure at this address want BOTH loans? -> ONE-OP.", 10, False),
 ("  If one of them looks defunct and the other active, that is still SEPARATE unless the", 10, False),
 ("  names tie them together.", 10, False), ("", 10, False),
 ("Do not look anything up. Yellow = unlabelled; when nothing is yellow you are done.", 11, True)]
for t, sz, b in L:
    rs.append([t]); rs.cell(row=rs.max_row, column=1).font = Font(size=sz, bold=b)
rs.column_dimensions["A"].width = 100
wb.move_sheet("SCOPE & RULES", offset=-1); wb.active = 1
wb.save(OUT)
print(f"wrote {OUT}  ({len(k)} rows, labels blank)")
print(k.groupby(["region", "question"]).size().to_string())
