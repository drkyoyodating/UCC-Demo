#!/usr/bin/env python
"""P6 -- build the labelling workbook: both batches, one file, dropdowns.

Tooling only. Produces docs/UCC_labelling.xlsx for import into Google Sheets.
The pipeline does not depend on this and openpyxl is in requirements-optional.

Blindness is preserved exactly as in the CSVs: names and addresses only. No
weight, no prediction, no cluster id, no stratum, no batch-purpose hint.
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "UCC_labelling.xlsx"

HDR = PatternFill("solid", fgColor="1F3864")
BAND = PatternFill("solid", fgColor="EEF2F8")
TODO = PatternFill("solid", fgColor="FFF3CD")
THIN = Side(style="thin", color="C8CDD6")


def main() -> int:
    a = pd.read_csv(ROOT / "docs" / "labels_blank.csv", dtype=str).fillna("")
    b = pd.read_csv(ROOT / "docs" / "labels_blank_batch2.csv", dtype=str).fillna("")
    df = pd.concat([a, b], ignore_index=True)
    print(f"batch1={len(a)}  batch2={len(b)}  total={len(df)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "LABEL THESE"
    cols = ["pair_id",
            "A — name", "A — address", "A — city", "A — state", "A — zip",
            "B — name", "B — address", "B — city", "B — state", "B — zip",
            "label", "note"]
    ws.append(cols)
    for i, c in enumerate(ws[1], 1):
        c.fill = HDR
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
    src = ["pair_id", "a_name", "a_address", "a_city", "a_state", "a_zip",
           "b_name", "b_address", "b_city", "b_state", "b_zip", "label", "note"]
    for _, r in df.iterrows():
        ws.append([r[c] for c in src])

    n = len(df) + 1
    # visually separate the A block from the B block so the eye compares the right things
    for row in range(2, n + 1):
        for col in range(7, 12):
            ws.cell(row=row, column=col).fill = BAND
        for col in range(1, 14):
            ws.cell(row=row, column=col).border = Border(bottom=THIN)

    dv = DataValidation(type="list", formula1='"SAME,DIFFERENT,UNSURE"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Pick SAME, DIFFERENT or UNSURE."
    dv.errorTitle = "Not a valid label"
    ws.add_data_validation(dv)
    dv.add(f"L2:L{n}")

    # unlabelled rows glow, so a missed row is impossible to leave behind
    ws.conditional_formatting.add(
        f"A2:M{n}", FormulaRule(formula=['$L2=""'], fill=TODO, stopIfTrue=False))

    for col, w in zip("ABCDEFGHIJKLM", [9, 34, 26, 15, 7, 11, 34, 26, 15, 7, 11, 13, 26]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:M{n}"

    rs = wb.create_sheet("THE RULE")
    lines = [
        ("Are these two records the SAME FIRM?", 14, True),
        ("Put SAME / DIFFERENT / UNSURE in the `label` column. Unlabelled rows are highlighted yellow.", 11, False),
        ("", 11, False),
        ("STEP 1 — clean the fields first (these always run before anything else)", 12, True),
        ("P1  Address formatting is not an address difference. PO BOX = P.O. BOX. E = EAST. CO RD = COUNTY RD.", 11, False),
        ("      1200 S. TOWNSEND = 1200 S TOWNSEND. A ZIP+4 is not a different ZIP.", 11, False),
        ("P2  A ZIP that disagrees while name and street agree is a DATA ERROR. Ignore it.", 11, False),
        ("P3  A wrong state code on an obviously-Colorado record is a DATA ERROR. Ignore it.", 11, False),
        ("P4  A blank field is NOT evidence. Decide on the fields that are present.", 11, False),
        ("", 11, False),
        ("STEP 2 — then apply these, in order. Lower number wins.", 12, True),
        ("R0  Identical name AND identical address  ->  SAME", 11, False),
        ("R1  Plainly different firms, no shared address  ->  DIFFERENT      <- the commonest case", 11, False),
        ("R2  One name is a typo / truncation / abbreviation of the other  ->  SAME", 11, False),
        ("      AMERICAN NATIONAL BANJ = AMERICAN NATIONAL BANK.  COOPERS CONST = COOPERS CONSTRUCTION.", 11, False),
        ("      COLO NATIONAL BANK = COLORADO NATIONAL BANK.  Tested BEFORE R6.", 11, False),
        ("R3  Suffix-only difference (LLC vs INC)  ->  SAME", 11, False),
        ("R4  Same name, same city, different address  ->  SAME   (moved, or a yard and an office)", 11, False),
        ("R5  Same name, DIFFERENT Colorado cities, no shared address  ->  DIFFERENT", 11, False),
        ("R6  DIFFERENT NAMES AT THE SAME ADDRESS  ->  DIFFERENT      <- THE IMPORTANT ONE", 11, True),
        ("      A shared address is NOT evidence of one firm. Registered agents, franchise HQs,", 11, False),
        ("      medical groups and family properties all park unrelated filers at one address.", 11, False),
        ("      ONE exception: numbered outlets of a chain (COUNTRY HARVEST BUFFET 103 vs 500) = SAME,", 11, False),
        ("      because the NAME says so, not the address.", 11, False),
        ("R7  Distinct legal entities of one family  ->  DIFFERENT  (WELLS FARGO BANK NA vs WF EQUIPMENT FINANCE INC)", 11, False),
        ("R8  Two DIFFERENT person-names at one address  ->  DIFFERENT  (spouses are different borrowers).", 11, False),
        ("      The SAME person at one address is SAME.", 11, False),
        ("R9  Successor / renamed entity  ->  DIFFERENT  (we label records as filed, not corporate history)", 11, False),
        ("R10 Undecidable after all of the above  ->  UNSURE, and move on.", 11, False),
        ("", 11, False),
        ("DO NOT look anything up. No Google, no Secretary of State search. The rule is the rule —", 12, True),
        ("outside research makes the labels unreproducible and is not available at scale.", 11, False),
    ]
    for txt, sz, bold in lines:
        rs.append([txt])
        rs.cell(row=rs.max_row, column=1).font = Font(size=sz, bold=bold)
    rs.column_dimensions["A"].width = 118
    wb.move_sheet("THE RULE", offset=-1)
    wb.active = 1
    wb.save(OUT)
    print(f"wrote {OUT}")
    for c in ("weight", "stratum", "N_h", "is_repeat", "cluster"):
        assert c not in [x.lower() for x in cols], c
    print("BLIND-FILE CHECK: no weight / stratum / N_h / cluster column. PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
